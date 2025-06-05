import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

DATA_DIR = "data"
EXCEL_PATH = os.path.join(DATA_DIR, "tasks.xlsx")

def save_tasks_to_excel(list_name, date_str, tasks_with_done_times):
    """
    Save tasks to Excel with spacing and done times.
    tasks_with_done_times: list of tuples (task_name, done_time_str)
    """
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)

    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"
    else:
        wb = load_workbook(EXCEL_PATH)
        if "Tasks" in wb.sheetnames:
            ws = wb["Tasks"]
        else:
            ws = wb.create_sheet("Tasks")

    # Add a blank row for spacing
    ws.append([])

    # Add list name row
    ws.append([f"{list_name} - {date_str}"])

    # Add each task with its done time (if any)
    for task in tasks_with_done_times:
        task_name = task[0]
        done_time = task[1] if len(task) > 1 else None

        if done_time:
            ws.append([f"{task_name} - {done_time}"])
        else:
            ws.append([task_name])

    wb.save(EXCEL_PATH)

def save_task_done_time(list_name, task_name, done_time=None):
    """
    Log the completion time of a task in the Excel file.
    It appends a new row with list name, task name, date, and completion time.
    done_time: a datetime object. If None, current time is used.
    """
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)

    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()
        ws = wb.active
        ws.title = "Task Completion Log"
        ws.append(["List Name", "Task Name", "Date", "Time Completed"])
    else:
        wb = load_workbook(EXCEL_PATH)
        if "Task Completion Log" in wb.sheetnames:
            ws = wb["Task Completion Log"]
        else:
            ws = wb.create_sheet("Task Completion Log")
            ws.append(["List Name", "Task Name", "Date", "Time Completed"])

    if done_time is None:
        done_time = datetime.now()
    date_str = done_time.strftime("%Y-%m-%d")
    time_str = done_time.strftime("%H:%M:%S")

    ws.append([list_name, task_name, date_str, time_str])

    wb.save(EXCEL_PATH)
